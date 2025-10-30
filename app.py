from flask import Flask, render_template, request, redirect, url_for
import mysql.connector
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
from werkzeug.utils import secure_filename
import fitz  # PyMuPDF
import base64

# --- Flask Setup ---
app = Flask(__name__)

UPLOAD_FOLDER = os.path.join('static', 'uploads')
PDF_FOLDER = os.path.join('static', 'pdfs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PDF_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# --- Database Setup ---
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="tamil",
    database="user_data"
)
cursor = db.cursor()

# --- Home Page ---
@app.route('/')
def home():
    return render_template('home.html')

# --- Form Page ---
@app.route('/form')
def form():
    return render_template('form.html')

# --- Submit Form ---
@app.route('/submit', methods=['POST'])
def submit():
    Register = request.form['Register']
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    department = request.form['department']
    category = request.form['category']

    requirements = request.form.getlist('requirements')
    requirements_str = ', '.join(requirements)

    # --- Photo Upload ---
    photo = request.files.get('photo')
    photo_path = ""
    if photo and photo.filename != '':
        safe_name = secure_filename(photo.filename)
        ext = os.path.splitext(safe_name)[1]
        category_folder = os.path.join(app.config['UPLOAD_FOLDER'], category)
        os.makedirs(category_folder, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename_only = f"{name}_{timestamp}{ext}"
        filepath = os.path.join(category_folder, filename_only)
        photo.save(filepath)
        photo_path = f"{category}/{filename_only}"

    # --- Digital Signature ---
    signature_data = request.form.get('signature_data')
    signature_path = ""
    if signature_data:
        try:
            signature_data = signature_data.split(',')[1]
            signature_bytes = base64.b64decode(signature_data)
            sig_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'signatures')
            os.makedirs(sig_folder, exist_ok=True)
            sig_filename = f"{name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
            sig_path = os.path.join(sig_folder, sig_filename)
            with open(sig_path, "wb") as f:
                f.write(signature_bytes)
            signature_path = f"signatures/{sig_filename}"
        except Exception as e:
            print("Signature save error:", e)

    # --- Insert into MySQL ---
    sql = """
    INSERT INTO tamil (Register, name, email, phone, department, category, requirements, photo, signature)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    values = (Register, name, email, phone, department, category, requirements_str, photo_path, signature_path)
    cursor.execute(sql, values)
    db.commit()

    # --- Save to Excel ---
    excel_path = 'form_submissions.xlsx'
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(["Register", "Name", "Email", "Phone", "Department", "Date of Register", "Category", "Requirements", "Photo", "Signature"])
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb["Submissions"]
    date_of_register = datetime.now().strftime('%Y-%m-%d')
    ws.append([Register, name, email, phone, department, date_of_register, category, requirements_str, photo_path, signature_path])
    wb.save(excel_path)

    # --- Generate PDF ---
    template_pdf_path = r'C:\Users\SEEDLAB\Desktop\Tamilarasan\project_update\consonform.pdf'
    if not os.path.exists(template_pdf_path):
        return "PDF template not found!", 500

    doc = fitz.open(template_pdf_path)
    page = doc[-1]
    date_of_register = datetime.now().strftime('%d-%b-%Y')
    info_text = f"Name: {name}\nEmail: {email}\nReg No: {Register}\nDate: {date_of_register}"

    page_width = page.rect.width
    page_height = page.rect.height
    text_box = fitz.Rect(page_width - 210, page_height - 120, page_width - 10, page_height - 20)
    page.insert_textbox(text_box, info_text, fontsize=12, fontname="helv", color=(0, 0, 0), align=1)

    signature_full_path = os.path.join(app.config['UPLOAD_FOLDER'], signature_path)
    if os.path.exists(signature_full_path):
        sig_rect = fitz.Rect(50, page_height - 150, 250, page_height - 50)
        page.insert_image(sig_rect, filename=signature_full_path)

    pdf_filename = f"{name}.pdf"
    pdf_output_path = os.path.join(PDF_FOLDER, pdf_filename)
    doc.save(pdf_output_path)
    doc.close()

    return redirect(url_for('success'))

# --- Success Page ---
@app.route('/success')
def success():
    return render_template('success.html')

# --- List Users ---
@app.route('/list')
def list_users():
    cursor.execute("SELECT id, Register, name, email, phone, department, DATE_FORMAT(date_of_Register, '%d-%b-%Y'), category, photo, signature FROM tamil")
    users = cursor.fetchall()
    return render_template('list.html', users=users)

# --- Photo Gallery ---
@app.route('/photo')
def photo_gallery():
    photos = []
    for root, dirs, files in os.walk(UPLOAD_FOLDER):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                rel_path = os.path.relpath(os.path.join(root, file), UPLOAD_FOLDER)
                if 'signatures' not in rel_path:
                    photos.append(rel_path)
    return render_template('photo.html', photos=photos)

# --- Contact Page ---
@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        message = request.form['message']
        print(f"Message received from {name} ({email}): {message}")
        return "<h2>Thank you for contacting us! We'll get back to you soon.</h2>"
    return render_template('contact.html')

# --- Run Server ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
